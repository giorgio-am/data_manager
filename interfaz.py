import openpyxl

import funciones_interfaz
import funciones_metas_sanitarias
import metas

try:
    import tkinter as tk                # python 3
    from tkinter import font as tkfont  # python 3
except ImportError:
    import Tkinter as tk     # python 2
    import tkFont as tkfont  # python 2
from procesos_dao import ProcesoDAO
from tkinter import ttk
from funciones_interfaz import *
from openpyxl import Workbook
from tkinter import filedialog
from functools import partial
import os
class SampleApp(tk.Tk):

    def __init__(self, *args, **kwargs):

        tk.Tk.__init__(self, *args, **kwargs)

        #self.title_font = tkfont.Font(family='Helvetica', size=18, weight="bold", slant="italic")
        big_frame = ttk.Frame(self)
        big_frame.pack(fill="both", expand=True)

        tema = 'azure.tcl'
        if '_MEIPASS2' in os.environ:
            tema = os.path.join(os.environ['_MEIPASS2'], tema)

        # Set the initial theme
        self.tk.call("source", tema)
        self.tk.call("set_theme", "light")

        def change_theme():
            # NOTE: The theme's real name is azure-<mode>
            if self.tk.call("ttk::style", "theme", "use") == "azure-dark":
                # Set light theme
                self.tk.call("set_theme", "light")
            else:
                # Set dark theme
                self.tk.call("set_theme", "dark")

        # Remember, you have to use ttk widgets
        button = ttk.Button(big_frame, text="Cambiar tema", command=change_theme)
        button.grid(row=0, column=0)
        # the container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = ttk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)


        self.frames = {}
        for F in (MenuPrincipal, Explorar, Administrar, Opciones, Kpi, MetasSanitarias, Ley18834, Ley19813 ):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("MenuPrincipal")

    def show_frame(self, page_name):
        '''Show a frame for the given page name'''
        frame = self.frames[page_name]
        frame.tkraise()


class MenuPrincipal(tk.Frame):

    def __init__(self, parent, controller, **kwargs):




        tk.Frame.__init__(self, parent, **kwargs)
        self.controller = controller
        #label = tk.Label(self, text="GESTIÓN POR PROCESOS SSBB", font=controller.title_font)

        #label.pack(side="top", fill="x", pady=10)
        #area_iz = ttk.Frame(width=400, height=660, bg="White", borderwidth=1, relief=RIDGE)

        frame_izquierda = tk.Frame(self, highlightbackground= "grey", highlightthickness = 1, width = 250, height = 400, bd = 0)
        frame_derecha = tk.Frame(self, highlightbackground= "grey", highlightthickness = 1, width = 250, height = 400, bd = 0)
        frame_derecha2 = tk.Frame(self, highlightbackground="grey", highlightthickness=1, width=250, height=400, bd=0)
        frame_derecha3 = tk.Frame(self, highlightbackground="grey", highlightthickness=1, width=250, height=400, bd=0)

        frame_izquierda.grid(rowspan=1, column=0, pady=20, padx=20)
        #frame_derecha.grid(row=0, column= 1)
        frame_derecha2.grid(row=0, column=2, pady=20, padx=20)
        frame_derecha3.grid(row=0, column=3, pady=20, padx=20)

        label = ttk.Label(frame_derecha2, text="      GPP SSBB       ")
        label.grid(row=1, column=1)
        label.config(font=("Default", 20))

        label_izq = ttk.Label(frame_izquierda, text="Instrumentos de Gestión")
        label_izq.grid(row=1, column=1)
        label_izq.config(font=("Default", 20))
        # label_columna_0 = ttk.Label(self, text=None)
        # label_columna_0.grid(row=5)
        # label_columna_1 = ttk.Label(self, text='                                      ')
        # label_columna_1.grid(column=0, rowspan=50)

        button1 = ttk.Button(frame_derecha2, text="Explorar",
                            command=lambda: controller.show_frame("Explorar"))
        button2 = ttk.Button(frame_derecha2, text="Administrar",
                            command=lambda: controller.show_frame("Administrar"))
        button3 = ttk.Button(frame_derecha2, text="Opciones",
                            command=lambda: controller.show_frame("Opciones"))

        button4 = ttk.Button(frame_izquierda, text="Metas Sanitarias",
                            command=lambda: controller.show_frame("MetasSanitarias"))

        button1.grid(row=2, column=1, sticky='nsew', pady=5, padx=10)
        button2.grid(row=3, column=1, sticky='nsew', pady=5, padx=10)
        button3.grid(row=4, column=1, sticky='nsew', pady=5, padx=10)

        button4.grid(row=2, column=1, sticky='nsew', pady=5, padx=10)

        label_der = ttk.Label(frame_derecha3, text="Gestión de Riesgos\n")
        label_der.grid(row=0, column=1)
        label_der.config(font=("Default", 20))

        button1_der = ttk.Button(frame_derecha3, text="Explorar")
        button1_der.grid(row=1, column=1, sticky='nsew', pady=5, padx=10)

class Explorar(tk.Frame):


    def __init__(self, parent, controller):
        ttk.Frame.__init__(self, parent)
        self.controller = controller

        nombre_proceso = tk.StringVar()
        subdirección = tk.StringVar()
        departamento = tk.StringVar()
        alcance = tk.StringVar()
        estado = tk.IntVar()

        filtrados = []

        subdirecciones = funciones_interfaz.listar_subdirecciones()
        dptos = funciones_interfaz.listar_dptos()
        nombres_dptos = []
        nombres_subdirecciones = []

        for dpto in dptos:
            nombre_dpto = dpto[0]
            nombres_dptos.append(nombre_dpto)

        for dpto in dptos:
            nombre_dpto = dpto[0]
            nombres_dptos.append(nombre_dpto)


        #dptos_2 = funciones_interfaz.listar_dptos()[1]
        frame_izquierda = ttk.Frame(self)
        frame_derecha = ttk.Frame(self)

        frame_izquierda.grid(row=1, column=0)
        frame_derecha.grid(row=1, column=1)

        #label = tk.Label(self, text="Explorar Procesos", font=controller.title_font)
        label = ttk.Label(self, text="Explorar Procesos", )
        label.config(font=("Default", 30))
        #label.pack(side="top", fill="x", pady=10)
        label.grid(row=0, columnspan=2)

        combo2 = ttk.Combobox(frame_izquierda, state='readonly', values=subdirecciones, textvariable=subdirección)
        combo = ttk.Combobox(frame_izquierda, state='readonly', values=nombres_dptos)



        #nombre_sd =



        tree = ttk.Treeview(frame_derecha, columns=(1, 2, 3, 4, 5, 6), height=20, show="headings")
        tree.grid(row=1, column=0, sticky='nsew', pady=5)

        tree.heading(1, text="Id")
        tree.heading(2, text="Proceso")
        tree.heading(3, text="Subdirección")
        tree.heading(4, text="Departamento")
        tree.heading(5, text="Alcance")
        tree.heading(6, text="Estado")

        tree.column(1, width=100)
        tree.column(2, width=100)
        tree.column(3, width=100)
        tree.column(4, width=150)
        tree.column(5, width=100)
        tree.column(6, width=100)
        # Inserting Scrollbar
        scroll = ttk.Scrollbar(frame_derecha, orient="vertical", command=tree.yview)
        scroll.grid(row=1, column=1, sticky='nse')

        tree.configure(yscrollcommand=scroll.set)

        # frame_derecha.grid_rowconfigure(0, weight=1)
        # frame_derecha.grid_columnconfigure(0, weight=1)

        #registros = funciones_interfaz.ver_todos_malo()

        def poblar():
            registros = funciones_interfaz.ver_todos_malo()
            combo.set('')
            combo2.set('')
            tree.delete(*tree.get_children())
            for registro in registros:
                tree.insert('', tk.END, values=registro[0:7])
        def volver():
            tree.delete(*tree.get_children())
            controller.show_frame("MenuPrincipal")

        def filtrar_dpto(event=None):
            registros = funciones_interfaz.ver_todos_malo()
            filtrada = []
            for registro in registros:
                if registro[3] == combo.get():
                    filtrada.append(registro)
            tree.delete(*tree.get_children())
            for filtro in filtrada:
                tree.insert("", tk.END, values=filtro[0:7])

        def filtrar_sd(event=None):
            registros = funciones_interfaz.ver_todos_malo()
            filtrada2 = []
            combo.set('')
            filtrados = []
            for dpto in dptos:
                #print(f'dpto[1]: {dpto[1]}')
                if dpto[1] == combo2.get():
                    filtrados.append(dpto[0])

            combo.config(values=filtrados)
            for registro in registros:
                if registro[2] == combo2.get():
                    filtrada2.append(registro)
            tree.delete(*tree.get_children())
            for filtro in filtrada2:
                tree.insert("", tk.END, values=filtro[0:7])

        def exportar():

            wb = Workbook()
            ws = wb.active
            files = [('All Files', '*.*'),
                     ('Excel files', '*.xlsx'),
                     ('Text Document', '*.txt')]

            archivo = filedialog.asksaveasfilename(title="Seleccionar Archivo", filetypes= files,  defaultextension= files)

            wb.save(filename=archivo)
            #workbook = openpyxl.load_workbook(filename=archivo)
            sheet = ws
            headers = ['Id', 'Proceso', 'Subdirección', 'Departamento', 'Alcance', 'Estado']
            sheet.delete_rows(idx=2, amount=15)
            sheet.append(headers)
            for row_id in tree.get_children():
                row = tree.item(row_id)['values']
                #print(row)
                sheet.append(row)
            wb.save(filename=archivo)
            win = Toplevel()
            win.title("Tabla Exportada")
            frame = ttk.Frame(win)
            frame.pack(expand=True)
            win.geometry("300x300")

            button_aceptar = ttk.Button(frame, text='Aceptar', command=win.destroy)
            win.grab_set()
            label_exito = ttk.Label(frame, text='Archivo creado con éxito')
            label_exito.pack()
            button_aceptar.pack()
            combo.set('')


        button = ttk.Button(frame_izquierda, text="Volver al Menú Principal",
                           command=volver)
        button2 = ttk.Button(frame_izquierda, text="          Ver Todos          ",
                            command=poblar)
        button3 = ttk.Button(frame_izquierda, text='Exportar', command=exportar)
        button4 = ttk.Button(frame_izquierda, text='Indicadores', command=lambda: controller.show_frame("Kpi"))

        button.pack(padx=20, pady=20)
        label_vacia_2 = ttk.Label(frame_izquierda, text=None)
        label_vacia_2.pack()
        button2.pack()
        label2 = ttk.Label(frame_izquierda, text="Filtrar por Subdirección")
        label2.pack(after=button2)

        combo.bind("<<ComboboxSelected>>", filtrar_dpto)
        #combo2.bind("<<ComboboxSelected>>", combo.val )
        #combo2.bind("<<ComboboxSelected>>", filtrar_sd)


        combo2.pack(after=label2)
        label3 = ttk.Label(frame_izquierda, text="Filtrar por Departamento")
        label3.pack(after=combo2)

        combo.pack(after=label3)
        combo2.bind("<<ComboboxSelected>>", filtrar_sd)

        label_vacia = ttk.Label(frame_izquierda, text=None)
        label_vacia.pack(after=combo)
        button3.pack(after=label_vacia, fill='x', padx=20)
        button4.pack(after=button3, fill='x', padx=20, pady=10)
        #def actualizar_dptos(event=None):
            # combo.config(show=None)
            # filtrados = []
            # for dpto in dptos:
            #     print(f'dpto[1]: {dpto[1]}')
            #     if dpto[1] == combo2.get():
            #         filtrados.append(dpto[0])
            #
            # combo.config(values=filtrados)


        #combo2.bind("<<ComboboxSelected>>", actualizar_dptos)

class Administrar(tk.Frame):

    def __init__(self, parent, controller):
        ttk.Frame.__init__(self, parent)
        self.controller = controller
        #label = tk.Label(self, text="Menú Administración", font=controller.title_font)
        label = ttk.Label(self, text="Menú Administración", )
        label.config(font=("Default", 30))
        label.pack(side="top", fill="x", pady=10, padx=250)
        button = ttk.Button(self, text="Volver al Menú Principal",
                           command=lambda: controller.show_frame("MenuPrincipal"))
        button2 = ttk.Button(self, text="Agregar Proceso",
                            command=nuevo_proceso)
        button3 = ttk.Button(self, text="Editar Proceso",
                            command=editar_proceso)
        button4 = ttk.Button(self, text="Eliminar Proceso",
                            command=eliminar_proceso)
        label_vacia = ttk.Label(self, text=None)

        button2.pack(pady=10)
        button3.pack(pady=10)
        button4.pack(pady=10)
        button.pack(pady=20)

        #label_vacia.pack(after=button2)

class Opciones(tk.Frame):

    def __init__(self, parent, controller):
        ttk.Frame.__init__(self, parent)
        self.controller = controller
        #label = tk.Label(self, text="Opciones", font=controller.title_font)
        label = ttk.Label(self, text="Opciones", )
        label.pack(side="top", fill="x", pady=10)
        button = ttk.Button(self, text="Volver al Menú Principal",
                           command=lambda: controller.show_frame("MenuPrincipal"))
        button.pack()

class Kpi(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        #label = tk.Label(self, text="Indicadores", font=controller.title_font)
        label = ttk.Label(self, text="Indicadores", )
        label.pack(side="top", fill="x", pady=10)
        button = ttk.Button(self, text="Atras",
                           command=lambda: controller.show_frame("Explorar"))
        button.pack()

class MetasSanitarias(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        #label = tk.Label(self, text="Indicadores", font=controller.title_font)
        label = ttk.Label(self, text="Metas Sanitarias", )
        label.pack(side="top", fill="x", pady=10)
        button_18 = ttk.Button(self, text="Ley 18.834",
                           command=lambda: controller.show_frame("Ley18834"))
        button_19 = ttk.Button(self, text="Ley 19.813",
                           command=lambda: controller.show_frame("Ley19813"))
        button = ttk.Button(self, text="Atras",
                           command=lambda: controller.show_frame("MenuPrincipal"))
        button_18.pack(pady=10)
        button_19.pack(pady=10)
        button.pack(pady=10)

class Ley18834(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        #label = tk.Label(self, text="Indicadores", font=controller.title_font)

        frame_izquierda = ttk.Frame(self)
        frame_derecha = ttk.Frame(self)

        frame_izquierda.grid(row=1, column=0)
        frame_derecha.grid(row=1, column=1)

        argumento = ''

        label = ttk.Label(frame_izquierda, text="Ley 18834", )
        label.pack(side="top", fill="x", pady=10)
        tree = ttk.Treeview(frame_derecha, columns=(1, 2, 3, 4, 5, 6), height=20, show="headings")
        label_ano = ttk.Label(frame_izquierda, text="Filtrar por Año")

        def poblar_metas():
            combo.set('')
            combo2.set('')
            resultados = funciones_metas_sanitarias.ver_todo()
            tree.delete(*tree.get_children())
            for resultado in resultados:
                tree.insert('', tk.END, values=resultado[0:7])


        año = StringVar()
        combo = ttk.Combobox(frame_izquierda, state='readonly', textvariable=año, values=('2011', '2012', '2013', '2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021'))
        establecimiento = StringVar()
        combo2 = ttk.Combobox(frame_izquierda, state='readonly',textvariable=establecimiento, values=('Complejo Asistencial Dr. Victor Rios Ruiz',
                                                                                                      'DSSBB', 'HFC Huepil','HFC Laja','HFC Mulchen',
                                                                                                      'HFC Nacimiento','HFC Santa Barbara','HFC Yumbel'))


        def filtro_ano(event=None):

            if combo2.get():
                tree.delete(*tree.get_children())
                registros = funciones_metas_sanitarias.ver_todo()
                filtrada = []
                for registro in registros:
                    if registro[1] == combo2.get() and registro[0] == int(combo.get()):
                        filtrada.append(registro)

                for filtro in filtrada:
                    tree.insert("", tk.END, values=filtro[0:7])
            else:
                tree.delete(*tree.get_children())
                registros = funciones_metas_sanitarias.ver_todo()
                filtrada = []
                for registro in registros:
                    if registro[0] == int(combo.get()):
                        filtrada.append(registro)

                for filtro in filtrada:
                    tree.insert("", tk.END, values=filtro[0:7])
            return combo.get()

        def filtro_establecimiento(event=None):
            if combo.get():
                tree.delete(*tree.get_children())
                registros = funciones_metas_sanitarias.ver_todo()
                filtrada = []
                for registro in registros:
                    if registro[1] == combo2.get() and registro[0] == int(combo.get()):
                        filtrada.append(registro)

                for filtro in filtrada:
                    tree.insert("", tk.END, values=filtro[0:7])
            else:
                tree.delete(*tree.get_children())
                registros = funciones_metas_sanitarias.ver_todo()
                filtrada = []
                for registro in registros:
                    if registro[1] == combo2.get():
                        filtrada.append(registro)

                for filtro in filtrada:
                    tree.insert("", tk.END, values=filtro[0:7])

        def saveExcel():

            wb = Workbook()
            ws = wb.active
            files = [('All Files', '*.*'),
                     ('Excel files', '*.xlsx'),
                     ('Text Document', '*.txt')]

            archivo = filedialog.asksaveasfilename(title="Seleccionar Archivo", filetypes= files,  defaultextension= files)
            if archivo:
                wb.save(filename=archivo)
                #workbook = openpyxl.load_workbook(filename=archivo)
                sheet = ws
                headers = ['Año', 'Establecimiento', 'Indicador', 'Meta', 'Resultado', 'Cumplimiento']
                sheet.delete_rows(idx=2, amount=15)
                sheet.append(headers)
                for row_id in tree.get_children():
                    row = tree.item(row_id)['values']
                    #print(row)
                    sheet.append(row)
                wb.save(filename=archivo)
                win = Toplevel()
                win.title("Tabla Exportada")
                frame = ttk.Frame(win)
                frame.pack(expand=True)
                win.geometry("300x300")

                button_aceptar = ttk.Button(frame, text='Aceptar', command=win.destroy)
                win.grab_set()
                label_exito = ttk.Label(frame, text='Archivo creado con éxito')
                label_exito.pack()
                button_aceptar.pack()
                combo.set('')


        def volver():
            combo.set('')
            tree.delete(*tree.get_children())
            controller.show_frame("MetasSanitarias")

        button_18 = ttk.Button(frame_izquierda, text="Ver Todo", command=poblar_metas
                           )
        button_19 = ttk.Button(frame_izquierda, text="Opciones",
                           )
        button = ttk.Button(frame_izquierda, text="Atras",
                           command=volver)

        button_exportar = ttk.Button(frame_izquierda, text='Exportar a Excel', command=saveExcel)

        #button_graficar = ttk.Button(frame_izquierda, text='Graficar', command= lambda arg=combo.get() : print(arg))

        button_18.pack(padx= 10, pady=10)
        button_19.pack(padx= 10, pady=10)
        button.pack(padx=10, pady=10)

        label_ano.pack(padx=10, pady=10)
        combo.pack(padx=10, pady=10)
        combo2.pack(padx=10, pady=10)

        button_exportar.pack()
        #button_graficar.pack()

        combo.bind("<<ComboboxSelected>>", filtro_ano)
        combo2.bind("<<ComboboxSelected>>", filtro_establecimiento)

        button_graficar = ttk.Button(frame_izquierda, text='Graficar',command= lambda forro = filtro_ano : metas.graficar_ano(forro()))

        def mostrar_graficar(event=None):
            button_graficar.pack(pady=10)

        combo.bind("<<ComboboxSelected>>", mostrar_graficar, add='+')


        tree.grid(row=1, column=0, sticky='nsew', pady=5)

        tree.heading(1, text="Año")
        tree.heading(2, text="Establecimiento")
        tree.heading(3, text="Indicador")
        tree.heading(4, text="Meta")
        tree.heading(5, text="Resultado")
        tree.heading(6, text="Cumplimiento")

        tree.column(1, width=50)
        tree.column(2, width=100)
        tree.column(3, width=280)
        tree.column(4, width=50)
        tree.column(5, width=60)
        tree.column(6, width=80)
        # Inserting Scrollbar
        scroll = ttk.Scrollbar(frame_derecha, orient="vertical", command=tree.yview)
        scroll.grid(row=1, column=1, sticky='nse')

        tree.configure(yscrollcommand=scroll.set)


class Ley19813(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        #label = tk.Label(self, text="Indicadores", font=controller.title_font)
        label = ttk.Label(self, text="Ley 19.813", )
        label.pack(side="top", fill="x", pady=10)
        button_18 = ttk.Button(self, text="Ver Todo",
                           )
        button_19 = ttk.Button(self, text="Opciones",
                           )
        button = ttk.Button(self, text="Atras",
                           command=lambda: controller.show_frame("MetasSanitarias"))
        button_18.pack(pady=10)
        button_19.pack(pady=10)
        button.pack(pady=10)



if __name__ == "__main__":
    app = SampleApp()
    app.mainloop()